from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os, time
def build_invoice_xlsx(path, invoice:dict):
    wb=Workbook(); ws=wb.active; ws.title='Invoice'
    headers=['invoice_id','order_id','user_id','card_name','buy_now','bought_for','fee_percent','variable_deduction','net_amount','currency','created_at','payment_status','provider_txn_id']
    ws.append(headers); ws.append([invoice.get(h) for h in headers])
    for i,h in enumerate(headers,1): ws.column_dimensions[get_column_letter(i)].width=max(12,len(h)+2)
    os.makedirs(os.path.dirname(path), exist_ok=True); wb.save(path); return path
